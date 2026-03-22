#!/usr/bin/env python3
"""Fill qrs-mawa-immo-verwaltung-angebot.xlsx with Immobilien-Software content.
Preserves xlsx2tex structure: sheet names, SUBSECTION, TABLESTART/TABLEEND, FIGURE, formulas."""

from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "qrs-mawa-immo-verwaltung-angebot.xlsx"


def main() -> None:
    wb = load_workbook(XLSX, data_only=False)

    # --- Cover Page ---
    cov = wb["Cover Page"]
    cov["B1"] = "logo.png"
    cov["B2"] = "mawacon.png"
    cov["B3"] = "Web-Anwendung zur Immobilienbewirtschaftung"
    cov["B4"] = "AG-26-IMMO-01"
    cov["B5"] = "23.02.2026"
    cov["B6"] = "QR Solutions GmbH"
    cov["B7"] = "Wilhelm-Hahn-Weg 29, 30457 Hannover"
    cov["B8"] = "USt-IdNr.: DE343396643"
    cov["B9"] = "mawa Immobilien"
    cov["B10"] = "Geschäftsführung"
    cov["B11"] = "nach Vereinbarung"
    cov["B12"] = "1.0"
    cov["B13"] = (
        "Dieses Dokument enthält ein Angebot inkl. Leistungsbeschreibung, "
        "Zeitplan und Glossar. Die Umsetzung erfolgt als moderne Browser-Anwendung "
        "mit sicherer Anmeldung und Speicherung der Daten auf Servern in der EU."
    )

    # --- Einleitung ---
    ein = wb["Einleitung"]
    ein["A1"] = (
        "Ziel ist die Entwicklung einer **Web-Anwendung** zur zentralen Verwaltung Ihrer "
        "Immobilien: *Kosten und Verbrauch*, *Verträge und Kündigungsfristen*, "
        "*Unterlagen* (z. B. Grundriss, Grundbuchauszug) sowie ein *Abgleich von Plan und Ist* "
        "auf Basis Ihrer Verträge und gebuchten Ausgaben. Die Lösung richtet sich an "
        "Eigentümer und Verwalter, die alle wichtigen Informationen pro Objekt an einem Ort bündeln möchten."
    )
    ein["A4"] = (
        "Die Umsetzung erfolgt in abgestimmten Projektphasen inklusive Tests und Übergabe. "
        "Als Grundlage dient ein **funktionaler Prototyp** (Demonstration), der die gewünschte "
        "Bedienlogik verdeutlicht; die Produktivversion erhält *persistente Datenbank*, "
        "*Benutzeranmeldung*, *Hosting* und *regelmäßige Backups* nach Absprache."
    )
    for _r in (2, 3, 5, 6, 7):
        ein[f"A{_r}"] = None
    ein["A8"] = "SUBSECTION"
    ein["B8"] = "Das Wichtigste in Kürze"
    ein["A9"] = "TABLESTART"
    ein["B9"] = "lX"
    ein["A10"] = "Thema"
    ein["B10"] = "Kurzfassung"
    ein["A11"] = "Ziel"
    ein["B11"] = (
        "Eine verständliche, sichere Anwendung für alle Immobilien mit Kostenüberblick, "
        "Vertragsdaten und Dokumenten – ohne Medienbrüche in Tabellenkalkulationen."
    )
    ein["A12"] = "Nutzen"
    ein["B12"] = (
        "Weniger Suchaufwand, bessere Transparenz bei laufenden Kosten, frühzeitige Erinnerung an "
        "Kündigungsfristen und nachvollziehbare Unterlagen für Verkauf, Vermietung oder Steuer."
    )
    ein["A13"] = "Dauer"
    ein["B13"] = "ca. 18–22 Wochen (ein Entwicklungsteam), abhängig von Freigaben und Hosting-Entscheidungen."
    ein["A14"] = "Voraussetzungen"
    ein["B14"] = (
        "Ansprechpartner beim Auftraggeber, Beispieldaten (Objekte, Kostenarten), "
        "Entscheidung zu Hosting (Auftraggeber oder Anbieter), Domain/Zugänge nach Absprache."
    )
    ein["A15"] = "Abnahme"
    ein["B15"] = (
        "An den vereinbarten Akzeptanzkriterien (siehe Anlage A); funktionale Tests "
        "und gemeinsame Abnahme mit Protokoll."
    )
    ein["A16"] = "Kosten"
    ein["B16"] = "Netto und Brutto siehe Blatt *Angebot* (Positionssumme inkl. 19 % USt)."
    ein["A17"] = "TABLEEND"

    # --- Angebot (keep formulas in F, B10-B12) ---
    ang = wb["Angebot"]
    rows = [
        (1, "Konzept & Architektur", "Ziele, Nutzerrollen, technische Architektur der Web-Anwendung", 12),
        (2, "Server, Datenbank & Sicherheit", "Speicherung in der EU, Anmeldung, Backups, Grundschutz", 28),
        (3, "Oberfläche Immobilien", "Objekte, Kosten, Verbrauch, Kategorien, Diagramme", 24),
        (
            4,
            "Verträge & Plan/Ist",
            "Vertragsdaten, monatliche/jährliche Kosten, Soll-Ist-Vergleich; optional Jahresverbrauch "
            "(Strom/Gas/Wasser) und optionale Ablage einer Vertrags-PDF ohne automatische Texterkennung",
            14,
        ),
        (5, "Dokumente & Belege", "Ablage Unterlagen, Upload von Belegen, Verknüpfung zu Buchungen", 10),
        (6, "Qualitätssicherung", "Tests, Fehlerbehebung, Abnahmevorbereitung", 18),
        (7, "Produktivgang", "Deployment, Monitoring-Grundlagen, Einrichtung Hosting", 12),
        (8, "Schulung & Übergabe", "Handbuch, Einweisung, Übergabeprotokoll", 20),
    ]
    for pos, leistung, details, stunden in rows:
        r = pos + 1
        ang.cell(r, 1, pos)
        ang.cell(r, 2, leistung)
        ang.cell(r, 3, details)
        ang.cell(r, 4, stunden)
        ang.cell(r, 5, 90)
        # F has formula =D*E

    # --- Abrechnung (keep TABLESTART/TABLEEND and formulas) ---
    abr = wb["Abrechnung"]
    abr["A3"] = "Projektstart (M1)"
    abr["B3"] = (
        "Auftragserteilung und Bereitstellung der Zugänge (Ansprechpartner, Beispieldaten, "
        "Hosting-/Domain-Entscheid); Kick-off."
    )
    abr["A4"] = "Zwischenrelease (M2)"
    abr["B4"] = (
        "Lieferung einer testbaren Version mit Kernfunktionen (Objekte, Kosten, Verträge); "
        "Feedbackschleife mit dem Auftraggeber."
    )
    abr["A5"] = "Abnahme & Produktivgang (M3)"
    abr["B5"] = (
        "Abnahme nach Akzeptanzkriterien, Produktivstellung, Einweisung und Übergabe der Dokumentation."
    )
    abr["C3"] = 0.3
    abr["C4"] = 0.4
    abr["C5"] = 0.3
    for cell in (abr["C3"], abr["C4"], abr["C5"]):
        cell.number_format = "0%"

    abr["A8"] = (
        "Gültigkeit: 30 Tage ab Angebotsdatum.\n"
        "Projektstart nach schriftlicher Beauftragung und Klärung der Hosting- und Zugangsfragen.\n"
        "Änderungen über dem beschriebenen Umfang werden als Change Request gesondert angeboten."
    )

    # --- Anlage A ---
    a = wb["Anlage A - Leistungen"]
    a["A1"] = "SUBSECTION"
    a["B1"] = "A1 Ziele"
    a["A2"] = "BULLETSTART"
    a["A3"] = (
        "Zentrale Übersicht über alle Immobilien mit laufenden Kosten, Verträgen und wichtigen Unterlagen."
    )
    a["A4"] = (
        "Transparenz bei Ausgaben und optionalen Zählerständen (Strom, Gas, Wasser) je Objekt."
    )
    a["A5"] = (
        "Unterstützung bei Fristen: Kündigungsfristen aus Verträgen sichtbar und erinnerbar."
    )
    a["A6"] = (
        "Plan vs. Ist: vertragliche Jahreskosten gegen gebuchte Ausgaben; bei Verbrauchskategorien optional "
        "Abgleich mit Zähler und im Vertrag hinterlegtem Jahresverbrauch – ohne automatische PDF-Auswertung."
    )
    a["A7"] = "BULLETEND"

    a["A9"] = "SUBSECTION"
    a["B9"] = "A2 Leistungsumfang"
    a["A10"] = "TABLESTART"
    a["B10"] = "llXr"
    packages = [
        ("L1", "Konzept & Architektur", "Workshops, Zielbild, technische Architektur der Web-App", "AK-01, AK-02"),
        ("L2", "Backend & Datenhaltung", "Datenbank, API, sichere Anmeldung, Backups", "AK-02, AK-03"),
        ("L3", "Oberfläche Nutzer", "Immobilien, Kosten, Kategorien, Verbrauch, Diagramme", "AK-04"),
        (
            "L4",
            "Verträge & Budgetlogik",
            "Verträge, monatlich/jährlich, Plan-Ist; optional Verbrauchswert und PDF je Vertrag",
            "AK-05",
        ),
        ("L5", "Dokumente", "Upload und Verwaltung von Unterlagen und Belegen", "AK-06"),
        ("L6", "Qualität & Abnahme", "Tests, Mängelbeseitigung, Abnahmeprotokoll", "AK-07, AK-08"),
        ("L7", "Produktivgang", "Deployment, Monitoring-Grundlagen", "AK-09"),
        ("L8", "Schulung & Dokumentation", "Handbuch, Schulungstermin, Übergabe", "AK-10"),
    ]
    r = 11
    for pid, name, det, ak in packages:
        a.cell(r, 1, pid)
        a.cell(r, 2, name)
        a.cell(r, 3, det)
        a.cell(r, 4, ak)
        r += 1
    a["A20"] = "TABLEEND"

    a["A22"] = "SUBSECTION"
    a["B22"] = "A3 Akzeptanzkriterien"
    a["A23"] = "TABLESTART"
    a["B23"] = "llXX"
    ak_rows = [
        (
            "AK-01",
            "L1",
            "Schriftlich freigegebenes Zielbild und Architekturkurzdokument liegen vor.",
            "Review-Meeting, freigegebenes PDF",
        ),
        (
            "AK-02",
            "L2",
            "Anmeldung nur für berechtigte Benutzer; Passwörter sicher gespeichert; Sitzung nach Timeout beendet.",
            "Testfälle Login/Logout",
        ),
        (
            "AK-03",
            "L2",
            "Daten werden in einer Datenbank gespeichert; tägliches Backup nach Konzept (nachweisbar).",
            "Backup-Report oder Screenshot",
        ),
        (
            "AK-04",
            "L3",
            "Mindestens ein Objekt anlegbar; Kosteneinträge und Kategorien sichtbar; Verbrauchsdaten optional erfassbar.",
            "Manueller Testlauf, Screenshots",
        ),
        (
            "AK-05",
            "L4",
            "Vertrag je Objekt erfassbar; Plan-Ist zeigt Soll aus Vertrag und Ist aus Buchungen je Kalenderjahr; "
            "optional Jahresverbrauch (Strom/Gas/Wasser) und optionale Vertrags-PDF speicherbar – ohne "
            "automatische Datenübernahme aus dem PDF.",
            "Stichprobe mit Demo-Daten",
        ),
        (
            "AK-06",
            "L5",
            "Datei-Upload für Unterlagen; mindestens ein Dokumenttyp (z. B. PDF) speicherbar und wieder auffindbar.",
            "Upload- und Download-Test",
        ),
        (
            "AK-07",
            "L6",
            "Definierte Testfälle werden ohne kritische Fehler durchlaufen.",
            "Testprotokoll",
        ),
        (
            "AK-08",
            "L6",
            "Abnahmeprotokoll mit Unterschrift (digital oder scanbar) liegt vor.",
            "Protokoll-PDF",
        ),
        (
            "AK-09",
            "L7",
            "Produktive URL erreichbar; Fehlerseiten und Logging nach Konzept.",
            "Smoke-Test Produktion",
        ),
        (
            "AK-10",
            "L8",
            "Kurzes Benutzerhandbuch (PDF) und ein Schulungstermin (online oder vor Ort) durchgeführt.",
            "Handbuch-Datei, Teilnehmerliste",
        ),
    ]
    r = 24
    for row_data in ak_rows:
        for c, val in enumerate(row_data, 1):
            a.cell(r, c, val)
        r += 1
    for clear_r in range(r, 37):
        for col in range(1, 5):
            a.cell(clear_r, col, None)
    a["A37"] = "TABLEEND"

    a["A39"] = "SUBSECTION"
    a["B39"] = "A4 Mitwirkungspflichten des Auftraggebers (mawa Immobilien)"
    a["A40"] = "BULLETSTART"
    a["A41"] = (
        "Benennung einer fachlichen Ansprechperson und eines Stellvertreters; Entscheidungen innerhalb von 5 Werktagen."
    )
    a["A42"] = (
        "Bereitstellung von Beispielobjekten, typischen Kostenarten und (falls vorhanden) anonymisierten Vertragsmustern."
    )
    a["A43"] = (
        "Klärung des Hostings: Server beim Auftraggeber, beim Anbieter oder gemietete Cloud (EU-Standort bevorzugt)."
    )
    a["A44"] = (
        "Bereitstellung von Zugängen (Domain, DNS, E-Mail für Benachrichtigungen falls gewünscht) nach technischer Vorgabe."
    )
    a["A45"] = (
        "Zeitnahes Feedback in den Review-Terminen; Verzögerungen seitens des Auftraggebers verlängern den Zeitplan entsprechend."
    )
    a["A46"] = "BULLETEND"

    a["A49"] = "SUBSECTION"
    a["B49"] = "A5 Betrieb & Abhängigkeiten"
    a["A50"] = "BULLETSTART"
    a["A51"] = (
        "Verfügbarkeit und Performance hängen von gewähltem Hosting, Internetanbindung und Wartung der Server ab."
    )
    a["A52"] = (
        "E-Mail-Benachrichtigungen setzen funktionierendes E-Mail-Routing beim Auftraggeber oder Provider voraus."
    )
    a["A53"] = (
        "Abnahme erfolgt nach den Kriterien in A3; reine Wunschliste ohne vertraglichen Bezug ist kein Abnahmehindernis."
    )
    a["A54"] = "BULLETEND"

    a["A56"] = "SUBSECTION"
    a["B56"] = "A6 Abgrenzung"
    a["A57"] = "TABLESTART"
    a["B57"] = "lXXX"
    a["A58"] = "Thema"
    a["B58"] = "Enthalten"
    a["C58"] = "Nicht enthalten"
    a["D58"] = "Wirkung bei Erweiterung"
    scope = [
        (
            "Funktionsumfang",
            "Leistungen laut Angebot und Anlage A (L1–L8)",
            "Beliebige Zusatzmodule ohne Beschreibung",
            "Change Request",
        ),
        (
            "Buchhaltung",
            "Erfassung und Auswertung von Kosten in der App",
            "DATEV-Export, doppelte Buchführung, Steuerberatung",
            "Nicht Gegenstand / gesondert",
        ),
        (
            "Belege",
            "Manueller Upload und Zuordnung",
            "Vollautomatische Texterkennung (OCR) aller Rechnungen",
            "Change Request",
        ),
        (
            "Recht",
            "Fachliche Unterstützung bei Anforderungen",
            "Rechtsberatung, Vertragsprüfung",
            "Nicht Gegenstand",
        ),
        (
            "Design",
            "Professionelles Standard-Layout (Corporate-Farben nach CI-Vorgabe wenn geliefert)",
            "Umfangreiches Redesign, Illustrationen, Marketing-Website",
            "Change Request",
        ),
        (
            "Mobile App",
            "Nutzbare Oberfläche im mobilen Browser",
            "Native iOS/Android-App in App-Stores",
            "Change Request",
        ),
        (
            "Support nach Übergabe",
            "Übergabe und vereinbarter Fixzeitraum (nach Absprache)",
            "24/7-Support ohne Vereinbarung",
            "Wartungsvertrag",
        ),
        (
            "Schulungen",
            "Eine Schulungseinheit inklusive",
            "Mehrfachschulungen für viele Standorte",
            "Change Request",
        ),
    ]
    r = 59
    for row_data in scope:
        for c, val in enumerate(row_data, 1):
            a.cell(r, c, val)
        r += 1
    a["A70"] = "TABLEEND"
    for clear_r in range(67, 70):
        for col in range(1, 6):
            a.cell(clear_r, col, None)

    # --- Anlage B ---
    b = wb["Anlage B - Zeitplan"]
    b["A1"] = "SUBSECTION"
    b["B1"] = "B1 Projektphasen"
    b["A2"] = "TABLESTART"
    b["B2"] = "lXlX"
    b["A3"] = "Phase"
    b["B3"] = "Meilenstein"
    b["C3"] = "Zeitraum"
    b["D3"] = "Kernergebnis / AK"
    phases = [
        ("P0 / M1", "Kick-off & Anforderungen", "W1–W2", "AK-01 freigegeben; Zugänge und Hosting geklärt"),
        ("P1", "Backend & Sicherheit", "W2–W6", "AK-02, AK-03 erfüllt"),
        ("P2", "Oberfläche & Objektdaten", "W6–W11", "AK-04 erfüllt"),
        ("P3 / M2", "Verträge, Dokumente, Plan/Ist", "W11–W14", "AK-05, AK-06 erfüllt"),
        ("P4", "Tests & Stabilisierung", "W14–W17", "AK-07 erfüllt"),
        ("P5", "Abnahme & Produktion", "W17–W19", "AK-08, AK-09 erfüllt"),
        ("P6 / M3", "Schulung & Übergabe", "W19–W22", "AK-10 erfüllt"),
    ]
    r = 4
    for row_data in phases:
        for c, val in enumerate(row_data, 1):
            b.cell(r, c, val)
        r += 1
    b["A11"] = "TABLEEND"
    for clear_r in range(12, b.max_row + 1):
        for col in range(1, 7):
            b.cell(clear_r, col, None)

    # --- Anlage C (FIGURE rows) ---
    c = wb["Anlage C - Prototyp"]
    captions = [
        "Übersicht – alle Immobilien",
        "Objekt auswählen – Kennzahlen",
        "Kosten nach Kategorien",
        "Diagramme Verbrauch und Kosten",
        "Kosteneinträge und Beleg-Upload",
        "Verträge und Kündigungsfristen (globale Übersicht)",
        "Verträge je Objekt (Tab inkl. optional PDF)",
        "Unterlagen (Grundriss, Grundbuch, …)",
        "Plan vs. Ist – Jahresvergleich (Kosten)",
        "Kategorie – Vertrag vs. Ist (Kosten; Verbrauch wenn hinterlegt)",
        "Neue Kategorie anlegen",
        "Dokument scannen / zuordnen",
        "Grenzwerte Verbrauch",
    ]
    for i, cap in enumerate(captions, start=1):
        c.cell(i, 1, "FIGURE")
        c.cell(i, 2, cap)
        c.cell(i, 3, "placeholder.svg")
        c.cell(i, 4, "0.85\\textwidth")
        c.cell(i, 5, None)

    # --- Anlage D - Glossar ---
    d = wb["Anlage D - Glossar"]
    d["A1"] = "TABLESTART"
    d["B1"] = "lX"
    d["A2"] = "Begriff"
    d["B2"] = "Erklärung"
    glossary = [
        ("Web-Anwendung", "Programm, das Sie im Browser nutzen – ohne Installation auf dem PC."),
        ("Hosting", "Bereitstellung der Server, auf denen die Anwendung und die Daten laufen."),
        ("Cloud / EU-Standort", "Daten werden in einem Rechenzentrum gespeichert; Standort EU reduziert datenschutzrechtliche Hürden."),
        ("Backup", "Regelmäßige Kopie der Datenbank zur Wiederherstellung bei Ausfall."),
        ("Anmeldung (Login)", "Geschützter Zugang nur für berechtigte Personen mit Benutzername und Passwort."),
        ("Rolle / Berechtigung", "Steuert, wer welche Funktionen sehen oder bearbeiten darf."),
        ("Objekt (Immobilie)", "Ein verwaltetes Gebäude oder eine Wohnung mit Adresse und Stammdaten."),
        ("Kosteneintrag", "Eine gebuchte Ausgabe oder Einnahme mit Betrag, Datum und Kategorie."),
        ("Kategorie", "Gruppierung von Kosten (z. B. Strom, Versicherung); kann erweitert werden."),
        ("Zählerstand", "Ablesung eines Verbrauchszählers (z. B. kWh, m³) zu einem Datum."),
        (
            "Vertrag",
            "Erfassung von Anbieter, Laufzeit und Kosten (monatlich oder jährlich); Grundlage für "
            "kostenbezogene Planwerte. Optional: erwarteter Jahresverbrauch bei Strom, Gas und Wasser; "
            "optional eine Vertrags-PDF zur Ablage. Keine automatische Auslesung von Beträgen oder Verbrauch aus dem PDF.",
        ),
        (
            "Jahresverbrauch (Vertrag)",
            "Optionaler Wert laut Vertrag in der passenden Einheit (z. B. kWh, m³) – als Orientierung für "
            "Vergleiche mit Zählerständen, nicht als Messersatz.",
        ),
        (
            "Vertrags-PDF",
            "Optionale Kopie des Vertragsdokuments zum Nachschlagen. Die Anwendung übernimmt daraus keine Daten automatisch.",
        ),
        (
            "Plan vs. Ist",
            "Vergleich vertraglich anzusetzender Jahreskosten mit gebuchten Ausgaben je Kalenderjahr; bei Strom/Gas/Wasser "
            "zusätzlich – sofern im Vertrag hinterlegt – Abgleich des Verbrauchs mit Zählerwerten im Jahr (vereinfachte Darstellung).",
        ),
        ("Kündigungsfrist", "Zeit, innerhalb der gekündigt werden muss, damit der Vertrag endet."),
        (
            "Grenzwert (Verbrauch)",
            "Selbst festgelegte Schwellen für den Jahresverbrauch; Überschreitungen werden in der Übersicht kenntlich gemacht.",
        ),
        ("Dokument / Unterlage", "Datei wie PDF oder Foto, z. B. Grundriss oder Rechnungskopie."),
        ("Upload", "Hochladen einer Datei aus Ihrem Computer in die Anwendung."),
        ("API", "Technische Schnittstelle zwischen Programmen – für spätere Erweiterungen (z. B. Import)."),
        ("Datenbank", "Strukturierte Speicherung aller Einträge (Objekte, Kosten, Verträge, Dateien)."),
        ("Deployment", "Übertragung einer fertigen Version auf den Produktivserver."),
        ("Abnahme", "Formelle Bestätigung, dass die vereinbarten Kriterien erfüllt sind."),
        ("Change Request (CR)", "Zusätzlicher Wunsch außerhalb des Angebots – wird gesondert bepreist."),
        ("GoBD", "Grundsätze zur ordnungsmäßigen Führung und Aufbewahrung von Büchern; relevant für Aufbewahrung von Belegen."),
        ("DSGVO", "Datenschutz-Grundverordnung; regelt den Umgang mit personenbezogenen Daten."),
        ("SSL/TLS", "Verschlüsselung der Datenübertragung zwischen Browser und Server (https)."),
        ("Staging", "Testumgebung, die der Produktion ähnelt, für Abnahme vor dem Livegang."),
        ("Produktivsystem", "Live-Umgebung, die Sie im Alltag nutzen."),
        ("UAT", "User Acceptance Test – fachliche Prüfung durch den Kunden vor Abnahme."),
        ("Handbuch", "Schriftliche Anleitung zur Bedienung der Anwendung."),
        ("SLA", "Service Level Agreement – vereinbarte Reaktionszeiten im Wartungsvertrag (optional)."),
        ("Wartung", "Fehlerbehebungen und kleine Updates nach Projektende – typisch als Monatsvertrag."),
        ("Responsive Design", "Die Seite passt sich Bildschirmgrößen an (PC, Tablet)."),
        ("Browser", "Programm wie Chrome, Edge oder Safari zum Aufrufen der Web-Anwendung."),
        (
            "Prototyp (Demonstration)",
            "Vorführung der Bedienlogik im Browser; Umfang und Speicherung können sich von der späteren Produktivversion unterscheiden.",
        ),
    ]
    r = 3
    for term, expl in glossary:
        d.cell(r, 1, term)
        d.cell(r, 2, expl)
        r += 1
    # TABLEEND row after last entry
    table_end_row = r
    d.cell(table_end_row, 1, "TABLEEND")
    d.cell(table_end_row, 2, None)
    for clear_r in range(table_end_row + 1, d.max_row + 1):
        for col in range(1, 6):
            d.cell(clear_r, col, None)

    wb.save(XLSX)
    print("Saved", XLSX)


if __name__ == "__main__":
    main()
